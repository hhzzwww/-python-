# -*- codeing = Utf-8 -*-
# @Time :2023/4/22 22:30
# @Author :ZunKki
# @File :python爬虫.py
# @Software: PyCharm

# import requests

# url = 'https://pvp.qq.com/web201605/js/herolist.json'

# response = requests.get(url)

# print(response.json())
"""获取数据的属性和方法"""
import re
import time

'''
print(response.text) #获取响应体字符串数据 -- str
print(response.content) #获取响应体二进制数据
print(response.json()) #获取响应体json数据，如果数据格式不是json数据格式，那么报错。
'''
'''查看响应体其他内容'''
'''
print(response.headers) #查看响应体的响应头信息
print(response.encoding) #指定响应体编码
print(response.apparent_encoding) #自动识别响应体编码

print(response.cookies) #获取响应体的cookies字段信息，得到的是RequestsCookieJar对象
print(response.cookies.get_dict()) #RequestsCookieJar对象转化为字典
print(response.url) #获取响应体的url地址
print(response.status_code) #获取响应体状态码
'''

'''
Origin:资源的起始位置
User-Agent：浏览器的身份标识
host：客户端制定想要访问服务器域名
Referer：告诉服务器，我是从那个页面过来的(防盗链)
cookies:用户的身份标识，常用来长时登录验证----能不加就不加，身份识别出来则进不去了


100 - 200  表示服务器成功接收到请求
200 - 299  表示请求成功
300 - 399  重定向（你需要请求的地址已经移动到了另一个资源位置）
400 - 499  发送请求的地址有错误（当前地址在服务器中没有数据）
500 - 599  服务器错误，服务器的问题
'''
# 状态码仅供参考，状态码是服务器返回的<逻辑>，如果服务器误导你，那么状态码仅供参考

'''
json数据
功能：目前前后端主流的数据交换格式
形式：外层{}、[]包裹，内层嵌套数据，规范的json数据用双引号
{字段1:值1,字段2:{嵌套字段1:嵌套值1,嵌套字段2:[{},{},{}]}}
和字典很像

在json数据中，字段值必须是以下的数据类型：
字符串
数字
对象（嵌套数据）
数组
布尔值
null #表示空

'''

'''
# json数据在用json()提取之前，类型是字符串
#用json()方法提取以后，会在方法底层经过数据类型的转换，转换称一个对象，为字典或者列表



'''

'''
method:请求方法 get post
url：请求网址(数据所在的地址)

headers：请求头字段的关键字参数
cookies：用户身份标识，字典
proxies：ip代理的关键字参数

params:(可选的)请求的查询参数
data:请求的请求参数<post请求>

allow_redirects:是否允许重定向，默认是允许的--就是通过各种方法将各种网络请求重新定个方向转到其它位置
timeout：设置请求响应的时间---时间超过会报错 
verify：是否验证证书<ca证书 ssl证书>   如果代码请求的地址没有证书，程序会报错
verify = False 忽视证书验证

json:json提交的请求参数<post请求>
auth：权限认证
'''

# import requests

# url = 'https://pic.sogou.com/napi/pc/searchList'
# for page in range(0, 144, 48):
#     params = {
#         "mode": "1",
#         "mood": "7",
#         "dm": "0",
#         "start": str(page),
#         "xml_len": "48",
#         "query": "风景"
#     }
#     response = requests.get(url=url, params=params)
#     json_data = response.json()  # 得到json数据--字典
#     # print(json_data)
#     data_list = json_data['data']['items']
#
#     for data in data_list:
#         img_url = data['locImageLink']
#         print(img_url)
#
# url = 'http://www.kfc.com.cn/kfccda/ashx/GetStoreList.ashx'
# params = {"op": "keyword"}
# for page in range(1, 17):
#     data = {
#         "cname": "",
#         "pid": "",
#         "keyword": "北京",
#         "pageIndex": str(page),
#         "pageSize": "10"
#     }  # 构建请求参数的关键字
#     # data 构建请求参数的关键字
#     response = requests.post(url=url, params=params, data=data, verify=False)
#     json_data = response.json()
#     # print(json_data)
#     data_list = json_data['Table1']
#     for data in data_list:
#         addressDetail = data['addressDetail']
#         cityName = data['cityName']
#         pro = data['pro']
#         print(addressDetail, cityName, pro, sep='|')

# 浏览器的地址导航栏只能发送get请求

'''
requests.exceptions.SSLError:网站没有证书引发的报错，因为requests模块会默认校验证书
解决办法：可以令verify = False 发送请求不校验证书
'''
'''
import requests

url = 'https://pic.sogou.com/napi/pc/searchList'
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
}
for page in range(0, 96, 48):
    params = {
        "mode": "1",
        "mood": "4",
        "dm": "0",
        "start": str(page),
        "xml_len": "48",
        "query": "风景"
    }
    response = requests.get(url=url, params=params, headers=headers).json()
    # print(response)
    data_list = response['data']['items']
    count = 1
    for data in data_list:
        img_url = data['picUrl']
        print(img_url)

        try:  # 异常捕获
            # 请求图片数据
            img_response = requests.get(url=img_url, timeout=3).content
            # 准备文件名
            file_name = str(count) + '.jpg'

            count += 1
            with open('img\\' + file_name, mode='wb') as f:
                f.write(img_response)
                print('保存完成：', file_name)
        except Exception as e:
            print(e)
    print('------------------------------------------------------\n')

'''
'''
爬虫项目的实现步骤
1.找数据所在的链接地址 <分析网页性质： 静态网页/动态网页>
2.代码模拟请求地址数据
3.数据提取，提取需要的数据，剔除不需要的数据
    css选择器：专门提取html数据
    正则表达式：万能的匹配字符串的方式，主要是在小范围的字符串中提取数据
    xpath节点提取：专门提取html的数据
4.数据保存（本地、数据库）
'''

# 所有的html数据其类型是字符串类型 response.text
# 假设 html 是我请求过来的数据
# 只有selector对象可以调用数据解析方法
'''
import parsel # parsel 数据解析模块, 第三方模块 pip install parsel

# 1.转换类型
selector = parsel.Selector(html)
print(selector)
# 2. 根据转化后的对象解析数据
result = selector.css('p')

result = selector.css('p').getall() #get() 从Selector 对象中提取第一个数据，直接返回字符串数据给我们
result = selector.css('p').get() # getall() 从Selector 对象中提取所有数据， 返回一个列表

# 在css语法中 空格代表取后代标签<子标签, 孙子标签, 重孙子标签 ....>
# 在css语法中 > 代表仅取子标签，不会取到下面两层

# . 代表提取标签的类型
# 具有相同类属性的标签都会被提取
# 类选择器可以通过标签的类属性（class属性）精确定位到你想要的标签
# 只有标签具有 class 属性才可以用类选择器提取对应的标签
# 如股票类属性中有空格， 需要将空格替换成 . 因为空格在css语法中有特殊含义

# '#' 使用id选择器提取数据     <class  > 用 '.' ；<id  > 用 '#'
result = selector.css('.top').getall()


# contend 代表标签的id属性
# 只要标签具有 id 属性， 那么所有具有相同id属性的标签都会被提取到
# id 在 html中一般是唯一的
result = selector.css('#contend').getall()


#组合选择器， 主要是加约束
# 一般标签选择器开头，后续的顺序自定义
result = selector.css('p.top#contend').getall()
print(result)

# ::属性选择器 当你提取到标签之后，需要对标签特定的值进行提取（标签包含的文本内容，标签的属性）
#text 提取标签包含的文本内容
result = selector.css('p.top#contend::text').getall()  

# ::attr(href) 根据标签中包含的属性名字提取属性值
# href 是属性名字
result = selector.css('a::attr(href)').getall()


 



import requests
import parsel

# 1.找数据地址
url = 'https://www.bqg78.com/book/1031/1.html'

# 2.发送请求
response = requests.get(url)
# print(response.text)

# 3.解析数据
# 3.1 转化对象
selector = parsel.Selector(response.text)  # 转化成selector 才能用css解释器
# print(selector)
# 3.2 提取数据
name_ = selector.css('div.content>h1').getall()
print(name_)

contend = selector.css('#chaptercontent').getall()
print(contend)

with open('a1.html', mode='w', encoding='utf-8') as f:
    f.write(response.text)
'''
'''
# 微医网数据抓取
import requests
import parsel

# 1.确认请求地址
url = 'https://www.wedoctor.com/search/expert?q=%E5%86%85%E7%A7%91'
# 2. 发送请求
response = requests.get(url)
html_data = response.text
# print(html_data)

# 3.数据解析
selector = parsel.Selector(html_data)
# print(selector)

# 数据第一次提取
lis = selector.css('.g-doctor-item')
# print(len(lis))

for li in lis:
    name = li.css(' .wrap>a::text').get()
    print(name)
    level = li.css(' dl dt::text').getall()[1].strip()
    kind = li.css('dl dd p:nth-child(1)::text').get()
    score = li.css('.star>em::text').get()
    good_for = li.css('.skill>p::text').get().strip().replace('\n', '').replace(' ', '')
    picture = li.css(' .infos.video span em:nth-child(1)::text').get()
    # print(picture)

# 当提取不到数据，可以从以下几个方面考虑
# 1.数据没有获取到 ，打印查看
# 2.可能是解析语法问题

headers = {
    'tcreq4log': '1',
    'r': '1683202717551',
    'logid': '10567339654217022913',
    'from': '0',
    'pu': 'sz%40224_220%2Cta%40iphone___24_112.0',
    'ct': '10',
    'cst': '1',
    'ref': 'index_iphone',
    'logFrom': 'rightTopBubble',
    'logInfo': '{"origin":"","from":"","channel":"","browserid":"24","qid":"10567339654217022913","timestamp":1683202717551}'
}


import requests
import parsel

url = 'https://m.maoyan.com/asgard/board/4'
headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
}

response = requests.get(url=url, headers=headers)
html_text = response.text
# print(html_text)

# print(response)
selector = parsel.Selector(html_text)
# print(select)
divs = selector.xpath('//div[@class="board-card clearfix"]')
# print(len(divs))

for div in divs:
    title = div.xpath('.//h3/text()').get()
    # print(title)
    actors = div.xpath('.//div[@class="actors"]/text()').get()
    # print(actors)
    data = div.xpath('.//div[@class="date"]/text()').get()
    # print(data)
    number = div.xpath('.//div/span/text()').get()
    # print(number)


import requests
import parsel

url = 'https://movie.douban.com/top250'
headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
}

response = requests.get(url=url, headers=headers)
# print(response.text)
html_data = response.text
selector = parsel.Selector(html_data)
# print(selector)
lis = selector.xpath('//*[@class="grid_view"]/li')
# print(len(lis))
# print(lis)
for li in lis:
    title = li.xpath('.//*[@class="title"]/text()').get().replace(' ', '').replace('\n', '')
    print(title)
    info = li.xpath('.//*[@class="bd"]/p[1]/text()').get().replace(' ', '').replace('\n', '')
    print(info)


'''

'''
爬虫项目的实现步骤

1.找数据所在的链接地址 <分析网页性质：静态网页/动态网页>
2.代码模拟请求地址数据
3.数据提取，提取需要的数据，剔除不需要的数据
4.保存数据（本地、数据库）


import openpyxl

# 1. 创建一个工作簿对象
work_book = openpyxl.Workbook()
print(work_book)
# 2. 创建表对象
sheet1 = work_book.create_sheet('表1')
# 数据填充要基于表格对象操作
# 方式一
# sheet1['C3'] = '哈哈'
# sheet1['A3'] = 'python'

# 方式二
# sheet1.cell(row=1, column=1).value = '11111111'
# sheet1.cell(row=2, column=2).value = '22222222'

# 方式三：在append()括号里面穿一个序列，那么会把这个序列作为一行写入到表格中去
sheet1.append([1, 2, 3, 4, 5])
sheet1.append((5, 6, 7, 8, 9))
sheet1.append({'A': '12345', 'B': 'abcde'})
# 3. 保存工作簿对象
work_book.save('实例.xlsx')
'''
'''
import openpyxl

# 根据excel文件路径读取表格
work = openpyxl.load_workbook('实例.xlsx')
# 获取工作簿对象里面有几张表格
print(work.sheetnames)

# 根据表名在工作簿中提取表格
sheet = work['Sheet']

print(sheet.max_row)  # 最大行
print(sheet.max_column)  # 最大列

# 获取第一行数据
for i in range(1, sheet.max_column + 1):
    for j in range(1, sheet.max_row + 1):
        print(sheet.cell(row=1, column=1).value)

'''
'''
Excel 保存数据
import requests
import parsel
import openpyxl

url = 'https://m.maoyan.com/asgard/board/4'
headers = {
    "Host": "m.maoyan.com",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
}

response = requests.get(url=url, headers=headers)
# print(response.text)
html_data = response.text
selector = parsel.Selector(html_data)
# print(html_data)
lis = selector.css('.board-card.clearfix')
# print(len(lis))
# 创建工作簿对象和表对象
work = openpyxl.Workbook()
sheet1 = work.active
for li in lis:
    name = li.css(' .title::text').getall()[0]
    # print(name)
    stars = li.css(' .actors::text').get()
    # print(stars)
    date = li.css(' .date::text').get()
    # print(date)
    score = li.css(' .number::text').get()
    # print(score)
    # print(name, stars, date, score, sep='|')
    sheet1.append([name, stars, date, score])

work.save('猫眼.xlsx')

'''

# 列表和字典数据可以直接转json格式

# {} []
'''
import json  # 内置模块

data = {
    'name': 'ACME',
    'shares': 100,
    'price': 542.23
}
'''
'''
json序列化操作：把对象转化成json字符串
'''
'''
json_str = json.dumps(data)
print(json_str)
print(type(json_str))
'''
'''
json反序列化操作：把json字符串转化成python对象
'''
'''
str1 = '["1", "2"]'
json_obj = json.loads(json_str)
json_obj1 = json.loads(str1)
print(json_obj)
print(type(json_obj))

import json

data = {
    'name': 'ACCE',
    'school': '青灯',
    'price': 5120
}

# ensure_ascii = False 不使用默认的unicode编码
json_str = json.dumps(data, ensure_ascii=False)
# print(type(json_str))

with open('data.json', mode='w', encoding='utf-8') as f:
    f.write(json_str)



# json数据
import requests
import parsel

url = 'https://m.maoyan.com/asgard/board/4'
headers = {
    "Host": "m.maoyan.com",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
}

response = requests.get(url=url, headers=headers)
# print(response.text)
html_data = response.text
selector = parsel.Selector(html_data)
# print(html_data)
lis = selector.css('.board-card.clearfix')
# print(len(lis))
json_list = []  # 定义一个空列表， 用于接受每一条数据
for li in lis:
    name = li.css(' .title::text').getall()[0]
    # print(name)
    stars = li.css(' .actors::text').get()
    # print(stars)
    date = li.css(' .date::text').get()
    # print(date)
    score = li.css(' .number::text').get()
    # print(score)
    # print(name, stars, date, score, sep='|')
    d = {'电影名称': name, '主演': stars, '上映时间': date, '评分': score}

    json_list.append(d)
# print(json_list)
# 序列化
json_str = json.dumps(json_list, ensure_ascii=False)
with open('猫眼.json', mode='w', encoding='utf-8') as f:
    f.write(json_str)

'''

'''
csv数据，默认是一行是一条数据
    数据中每个字段与字段之间用逗号分隔


import csv  # 内置

data = [[1, 2, 3, 4], [1, 2, 3, 4], [5, 6, 7, 8], [6, 7, 8, 9]]

with open('data.csv', mode='a', encoding='utf-8',newline='') as f:
    # newline = ''  指定数据新行是一个空字符串，不然会有数据空行
    # 实例化一个csv文件的对象，括号内部传打开的文件对象
    # writerow(i) 给一个列表， 会把列表中的数据按照行的方式写入csv文件中
    csv_write = csv.writer(f)
    for i in data:
        csv_write.writerow(i)



import requests
import parsel
import csv

url = 'https://m.maoyan.com/asgard/board/4'
headers = {
    "Host": "m.maoyan.com",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
}

response = requests.get(url=url, headers=headers)
# print(response.text)
html_data = response.text
selector = parsel.Selector(html_data)
# print(html_data)
lis = selector.css('.board-card.clearfix')
# print(len(lis))

with open('猫眼-列表.csv', mode='a', encoding='utf-8', newline='') as f:
    csv_write = csv.writer(f)
    f.write('电影名字,主演,上映时间,评分\n')  # 用普通文本写表头
    for li in lis:
        name = li.css(' .title::text').getall()[0]
        # print(name)
        stars = li.css(' .actors::text').get()
        # print(stars)
        date = li.css(' .date::text').get()
        # print(date)
        score = li.css(' .number::text').get()
        # print(score)
        # print(name, stars, date, score, sep='|')
        csv_write.writerow([name, stars, date, score])

'''

'''
selenium 初学

1.在浏览器的设置页面找到 浏览器 的版本号
2.在 https://registry.npmmirror.com/binary.html?path=chromedriver/ 页面找到对应版本的驱动下载
3.配置selenium
    3.1 可以配置到解释器中，作为全局配置，后续如果浏览器更新了，那么要主要在解释器中换掉老版本
    3.2 放在项目文件夹下

'''

"""
from selenium import webdriver  # 浏览器功能

# 1. 实例化浏览器对象
driver = webdriver.Chrome()

# 2. 进行浏览器的自动化
driver.get('https://www.baidu.com/')
# driver.save_screenshot('百度.png')  # 制定路径对浏览器页面截屏

# page_source 查看当前浏览器渲染的数据
# 这个数据可能和真实浏览器得到的数据不一样
print(driver.page_source)

# get_cookies() 获取页面请求后的cookie
# 用selenium模拟登录，可以获取到用户登陆后的cookies，这个cookies也可以对接到 requests

# 查看当前页面的url地址
print(driver.current_url)

# 最大化浏览器
# driver.maximize_window()

# 最小化浏览器
# driver.minimize_window()

# with open('a.html', mode='w', encoding='utf-8') as f:
#     f.write(driver.page_source)

# 3. 退出整个浏览器
driver.quit()
"""
'''
selenium做的所有自动化操作，都是基于需求去操作的
咱们平常用户是怎么操作浏览器的，那么咱们的代码和用户操作的顺序大致一致
浏览器页面的操作顺序，就决定了代码顺序
'''

# # 元素提取
# from selenium import webdriver
# from selenium.webdriver.common.by import By  # 定位器功能
#
# driver = webdriver.Chrome()
# driver.get('https:/www.douban.com/')
#
# '''
# find_element    提取符合条件的第一个标签对象，返回对象
# find_elements   提取符合条件的第一个标签对象，返回列表
# '''
# # 数据标签对象提取
#
# # 根据id属性值获取元素
# result = driver.find_element(By.ID, 'anony-reg-new')
# # print(result)
#
# # 根据标签的name属性
# result1 = driver.find_element(By.NAME, 'google-site-verification')
# # print(result1)
#
# # 根据标签的 class 属性值获取标签
# result2 = driver.find_element(By.CLASS_NAME, 'wrapper')
# # print(result2)
#
# # 根据标签包含的文本内容获取标签(精确匹配)
# result3 = driver.find_element(By.LINK_TEXT, '关于豆瓣')
# # print(result3)
#
# # 根据标签包含的文本内容获取标签(模糊匹配)
# result4 = driver.find_elements(By.PARTIAL_LINK_TEXT, '豆瓣')
# # print(result4)
# # print(len(result4))
#
# # 根据标签名字做定位
# result5 = driver.find_elements(By.TAG_NAME, 'div')
#
# # 根据 xpath语法做定位
#
# # 根据css语法做定位,在selenium中css语法定位不支持属性提取器（::text）
# result6 = driver.find_element(By.CSS_SELECTOR, '#anony-reg-new .app>a')
# # print(result6)
#
# # text 属性 可以根据标签对象提取包含的文本内容，支持链式调用
# result7 = driver.find_element(By.CSS_SELECTOR, '#anony-reg-new .app>a').text
# # print(result7)
#
# '''
# get_attribute('href')
#     方法. 根据标签对象的属性名，提取其属性值，支持链式调用
# '''
# result8 = driver.find_element(By.CSS_SELECTOR, '#anony-reg-new .app>a').get_attribute('href')
# print(result8)
#
# '''
# send_keys('流浪地球2')
#     基于获取到的标签对象输入数据,必须这个标签对象要有输入框
# '''
# driver.find_element(By.CSS_SELECTOR, '.inp>input').send_keys('流浪地球2')
#
# '''
# .click()    执行标签对象的点击操作,需要该标签对象具有点击事件
# '''
# driver.find_element(By.CSS_SELECTOR, '.bn>input').click()
#
# input()
# driver.quit()

# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By
#
# driver = webdriver.Chrome()
# driver.get('https://www.jd.com/')
# # 隐式等待,是一种智能化等待,虽然我设置了10秒等待时间,但是如果在这个事件之前页面渲染完了,那么不会死等
# # 如果网站的连通性不好,超过时间会报错
# # 基于浏览器对象设置了一个隐式等待,那么后续所有打开的页面都遵循这个规则
# driver.implicitly_wait(10)
# # print(driver.page_source)
#
# # 死等,页面滚动等待页面渲染
# # time.sleep(10)
#
#
# # driver.back()  # 后退到上一级页面
# # driver.forward()  # 前进到下一个页面
# # driver.refresh() # 页面刷新,在一个网页中,如果页面后退,会导致旧页面元素标签过期,不可用
#
#
#
#
# input()
# driver.quit()

'''
import time
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
driver.get('https://gitee.com/')
driver.implicitly_wait(10)
driver.maximize_window()

点击登录按钮
driver.find_element(By.CSS_SELECTOR, '.mb-0.contents>li>a').click()

输入用户信息
driver.find_element(By.CSS_SELECTOR, '.field>input:nth-child(1)').send_keys('15171286980')
time.sleep(2)

driver.find_element(By.CSS_SELECTOR, '#user_password').send_keys('hzw15171286980')
time.sleep(2)

点击用户登录
driver.find_element(By.NAME, 'commit').click()

print(driver.get_cookies())  # 获取登录后的cookie

# 点击豆瓣读书
# driver.find_element(By.CSS_SELECTOR, '.anony-nav-links>ul>li>a').click()

# driver.window_handles 获取当前浏览器所有的页面窗口
input()
driver.quit()
'''

'''
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()

driver.get('https://music.163.com/#/song?id=1450083773')

# print(driver.page_source)
'''
'''
默认情况下selenium是无法找到嵌套网页的数据
方式一:根据索引切换进入到嵌套网页
switch_to.frame() 切换嵌套网页
frame(0) 根据嵌套网页的索引切换进入,进入第一个嵌套网页
#driver.switch_to.frame(0)
print(driver.page_source)

方式二:根据嵌套网页的<iframe>对象,切换进入到嵌套网页
iframe = driver.find_element(By.CSS_SELECTOR, '#g_iframe')
driver.switch_to.frame(iframe)
print(driver.page_source)


退出嵌套网页
driver.switch_to.parent_frame()
'''
'''
iframe = driver.find_element(By.CSS_SELECTOR, '#g_iframe')
driver.switch_to.frame(iframe)
driver.switch_to.parent_frame()  # 从嵌套网页,切换到父级网页
print(driver.page_source)

input()
driver.quit()

'''
'''
# 02.下拉滚动条
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()

driver.get('https://www.douban.com/')

# document.documentElement.scrollTop    指定滚动条的位置
# document.documentElement.scrollTop = document.documentElement.scrollHeight    获取当前页面的最大高度
js = 'document.documentElement.scrollTop =500'

js_all = 'document.documentElement.scrollTop = document.documentElement.scrollHeight'

# 有些时候页面的高度会不断地改变， 需要给一定的时间去让浏览器渲染数据

# 通过浏览器对象执行js代码
driver.execute_script(js_all)

input()
driver.quit()

'''
'''
from selenium import webdriver
from selenium.webdriver.common.by import By
import csv


def get_product(keyword):
    # """指定关键字搜索商品数据"""
    # 找到搜索框，输入关键字
    driver.find_element(By.CSS_SELECTOR, '#key').send_keys(keyword)
    time.sleep(1)
    # 找到搜索按钮点击
    driver.find_element(By.CSS_SELECTOR, '.button').click()


def drop_down():
    """模拟人去滚动页面"""
    for h in range(1, 10, 2):  # 13579
        j = h / 9  # 1/9、3/9、5/9、7/9、9/9
        js_all = f'document.documentElement.scrollTop = document.documentElement.scrollHeight * {j}'
        driver.execute_script(js_all)
        time.sleep(0.5)  # 等待数据渲染


def parse_data():
    """解析数据函数"""
    lis = driver.find_elements(By.CSS_SELECTOR, '.gl-item')
    print(lis)
    for li in lis:
        title = li.find_element(By.CSS_SELECTOR, '.p-name.p-name-type-2>a>em').text
        title1 = title.replace('京品电脑', '').replace('\n', '')
        # print(title1)
        price = li.find_element(By.CSS_SELECTOR, '.p-price>strong>i').text
        # print(price)
        commit = li.find_element(By.CSS_SELECTOR, ' .p-commit>strong>a').text
        # print(commit)
        print(title1, price, commit, sep='|')

        with open('jingdong.csv', mode='a', encoding='utf-8', newline='') as f:
            csv_write = csv.writer(f)
            csv_write.writerow([title, price, commit])


def click_next():
    """翻页函数"""
    driver.find_element(By.CSS_SELECTOR, '.pn-next>em').click()


if __name__ == '__main__':
    word = input('请输入您要搜索商品的关键字：')
    driver = webdriver.Chrome()
    driver.get('https://www.jd.com/')
    driver.implicitly_wait(10)
    driver.maximize_window()

    # 调用搜索商品的函数
    get_product(word)

    for i in range(10):
        # 调用滚动页面的函数
        drop_down()
        # 数据解析
        parse_data()
        # 点击下一页
        time.sleep(3)
        click_next()

    input()
    driver.quit()
'''

'''
鼠标动作链
'''

# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver import ActionChains  # 导入鼠标动作链功能
#
# driver = webdriver.Chrome()
# driver.get('https://www.runoob.com/try/try.php?filename=jqueryui-api-droppable')
# driver.switch_to.frame(0)
#
# # 找到可以拖动的标签
# drag = driver.find_element(By.CSS_SELECTOR, '#draggable')
# # 找到放置的位置
# drop = driver.find_element(By.CSS_SELECTOR, '#droppable')
#
# """鼠标动作链操作，支持链式调用"""
# # 实例化一个动作链对象，括号内部需要传递当前driver浏览器对象
# action = ActionChains(driver)
#
# # 定义一个鼠标动作，但是动作到目前为止还没有执行，预编译
# action.drag_and_drop(drag, drop)
#
# # perform() 执行鼠标动作链
# action.perform()
#
# input()
# driver.quit()

'''
案例-码云登录后鼠标动作链解决消息提示框
'''

'''
9.无头模式，不需要打开浏览器，selenium的无头模式一般用于项目写完后添加，因为些项目我们需要看到浏览器运行效果
'''
# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options  # 谷歌浏览器的配置项功能
#
# # 声明一个谷歌配置对象
# option = Options()
#
# # 添加浏览器的无头模式
# option.add_argument('--headless')  # 添加无头模式
#
# driver = webdriver.Chrome(options=option)
# driver.get('https://gitee.com/')
# driver.implicitly_wait(10)
# driver.maximize_window()
#
# driver.find_element(By.CSS_SELECTOR, '.mb-0.contents>li>a').click()
#
# driver.find_element(By.CSS_SELECTOR, '.field>input:nth-child(1)').send_keys('15171286980')
# time.sleep(2)
#
# driver.find_element(By.CSS_SELECTOR, '#user_password').send_keys('hzw15171286980')
# time.sleep(2)
#
# driver.find_element(By.NAME, 'commit').click()
#
# input()
# driver.quit()
# import requests
#
# url = 'https://pic.sogou.com/napi/pc/searchList'
# headers = {
#     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
# }
# for page in range(0, 96, 48):
#     params = {
#         "mode": "1",
#         "mood": "4",
#         "dm": "0",
#         "start": str(page),
#         "xml_len": "48",
#         "query": "风景"
#     }
#     response = requests.get(url=url, params=params, headers=headers).json()
#     # print(response)
#     data_list = response['data']['items']
#     count = 1
#     for data in data_list:
#         img_url = data['picUrl']
#         print(img_url)

# try:  # 异常捕获
#     # 请求图片数据
#     img_response = requests.get(url=img_url, timeout=3).content
#     print(type(img_response))
#     # 准备文件名
#     file_name = str(count) + '.jpg'
#
#     count += 1
#     with open('img\\' + file_name, mode='wb') as f:
#         f.write(img_response)
#         print('保存完成：', file_name)
# except Exception as e:
#     print(e)
# print('------------------------------------------------------\n')

'''
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()

# 一定要放在实例化浏览器下面
# 修改 window.navigator.webdriver 属性值 === 注入js代码
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
        Object.defineProperty(navigator, 'webdriver', {
            get: () => False
        })
    """
})
driver.get('https://qikan.cqvip.com/index.html')
print(driver.page_source)

input()
driver.quit()
'''
'''
# 爬取网易云评论
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
driver.get('https://music.163.com/#/playlist?id=924680166')
driver.maximize_window()


def parse_data(): 
    """页面中提取数据"""
    divs = driver.find_elements(By.CSS_SELECTOR, '.itm')
    for div in divs:
        contend = div.find_element(By.CSS_SELECTOR, '.cnt.f-brk').text  # 获得名字加内容
        # print(contend)
        contend = re.findall('：(.*)', contend)[0]
        print(contend)
        with open('contend.txt',mode='a',encoding='utf-8') as f:
            f.write(contend + '\n')


# 进入iframe嵌套网页
driver.switch_to.frame(0)
# print(driver.page_source)

# 下拉页面
js = 'document.documentElement.scrollTop = document.documentElement.scrollHeight'
driver.execute_script(js)

for i in range(5):
    parse_data()

    # 解析下一页标签点击
    driver.find_element(By.CSS_SELECTOR, '.zbtn.znxt').click()
    time.sleep(2)

input()
driver.quit()
'''

'''
线程的逻辑执行
'''
#
# import threading
#
#
# def sing():
#     for i in range(3):
#         print('正在唱歌...%d' % i)
#         time.sleep(1)  # 模拟阻塞
#
#
# def dance():
#     for i in range(3):
#         print('正在跳舞...%d' % i)
#         time.sleep(1)
#
#         print('程序花费时间：', time.time() - start_time)
#
#
# start_time = time.time()  # 记录程序执行的开始时间
# # 所有的多任务转化对象都是基于函数对象做转化的
# sing_thread = threading.Thread(target=sing)
# # print(sing_thread)
# # 执行线程任务
# sing_thread.start()
#
# threading.Thread(target=dance).start()

'''
import threading
import time


def get(url, headers=None):
    print(url)
    time.sleep(5)
    print(headers)

  
urls = ['https://www.baidu.com', 'https://www.360.com', 'https://www.sousou.com']
for i in urls:
    threading.Thread(
        target=get,
        args=(i,),
        kwargs={'headers': {'user-agent': 'python-requests'}}
    ).start()

'''
'''
线程的参数传递
target  传入普通的函数对象只需要传递函数对象的名字，不要加括号
args    位置参数，一定传元组(如果只有一个参数，这个参数后面需要加逗号) **坑**
kwargs  传递关键字参数
'''

'''
多线程微医网数据抓取
'''

import requests
import parsel

import threading
import concurrent.futures

import csv

lock = threading.Lock()


def send_request(url):
    """发送请求的函数"""
    response = requests.get(url)
    return response.text


def parse_data(data):
    """解析数据的方法，传入数据进行解析"""
    selector = parsel.Selector(data)

    # 数据第一次提取
    lis = selector.css('.g-doctor-item')
    # print(len(lis))

    data_list = []  # 定义一个空列表，用于收集每一条数据
    for li in lis:
        name = li.css('.wrap>a::text').get()
        level = li.css('dl dt::text').getall()[1].strip()
        kind = li.css('dl dd p:nth-child(1)::text').get()
        Belonging = li.css('dl dd p:nth-child(2)>span::text').get()
        score = li.css('.star>em::text').get()
        good_for = li.css('.skill>p::text').get().strip().replace('\n', '').replace(' ', '')
        pic_see_price = li.css('.infos.image>span>em:nth-child(2)::text').get().strip()
        video_see_price = li.css('.infos.video>span>em:nth-child(2)::text').get().strip()
        print(name, level, kind, Belonging, score, good_for, pic_see_price, video_see_price)
        data_list.append([name, level, kind, Belonging, score, good_for, video_see_price])

    return data_list  # 返回嵌套列表，嵌套的是一条一条数据


def save_data(data_list):
    for data in data_list:
        # lock.acquire()  # 上锁
        # with open('微医网.csv', mode='a', encoding='utf-8', newline='') as f:
        #     csv_write = csv.writer(f)
        #     csv_write.writerow(data)
        # lock.release()  # 解锁
        with lock:
            with open('微医网.csv', mode='a', encoding='utf-8', newline='') as f:
                csv_write = csv.writer(f)
                csv_write.writerow(data)


def main(url):
    """主函数：调度其他三个函数的执行"""
    # 1.调用发送请求的函数
    html_data = send_request(url)
    # 2.调用解析数据的函数
    data_list = parse_data(html_data)  # 嵌套列表
    # 3.调用保存数据的方法
    save_data(data_list)
    # print('程序花费时间：', time.time() - start_time)


# main('https://www.wedoctor.com/expert/61409/%E5%86%85%E7%A7%91')

if __name__ == '__main__':
    # start_time = time.time()  # 记录程序执行的开始时间
    # for page in range(1, 39):
    #     url = f'https://www.wedoctor.com/expert/61409/%E5%86%85%E7%A7%91/p{page}'
    #
    #     threading.Thread(target=main, args=(url,)).start()

    # 线程池模块，防止直接采用多线程爬取导致的窗口过多而适得其反
    start_time = time.time()  # 记录程序执行的开始时间
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:  # 限制了最大任务数为3，所以相比直接多线程爬取所有数据要慢一点
        for page in range(1, 39):
            url = f'https://www.wedoctor.com/expert/61409/%E5%86%85%E7%A7%91/p{page}'
            executor.submit(main, url)

    print('线程池执行时间：', time.time() - start_time)











