# Author:Eric
# -*- codeing = utf-8 -*-
# @Time:2022/1/15 21:16
# @Author:86151
# @Site:
# @File:2.爬虫课.py
# # @Software:PyCharm
#
#
# import requests
# import re  # 正则表达式模块 ---帮助解析数据
#
# response = requests.get('https://www.xbiquge.la/23/23811/11700317.html')  # 字节型数据
# # print(response)
# # print(type(response))
# # encoding 表示设置对象的编码  apparent_encoding 自动识别响应体的编码
# response.encoding = response.apparent_encoding
#
# text = response.text
# # print(type(response.text))#字符串型
# # print(response.text)
# '''
# <meta name="keywords" content="(.*?)" />
#
# '''
# result = re.findall('<meta name="keywords" content="(.*?)" />', text, re.S)  # 数据解析以终端打印的数据为准
# result = result[0].strip()
# with open('a.txt', mode='a', encoding='utf-8') as f:
#     f.write(result)

"""

"""

'''
Origin:资源的起始位置
User-Agent：浏览器的身份标识
host：客户端制定想要访问服务器域名
Referer：告诉服务器，我是从那个页面过来的(防盗链)
cookies:用户的身份标识，常用来长时登录验证----能不加就不加，身份识别出来则进不去了


100 - 200  表示服务器成功接收到请求
200 - 299  表示请求成功
300 - 399  重定向（你需要请求的地址已经移动到了另一个资源位置）
400 - 499  发送请求的地址有错误（当前地址在服务器中没有数据）
500 - 599  服务器错误，服务器的问题
'''
# 状态码仅供参考，状态码是服务器返回的<逻辑>，如果服务器误导你，那么状态码仅供参考

'''
json数据
功能：目前前后端主流的数据交换格式
形式：外层{}、[]包裹，内层嵌套数据，规范的json数据用双引号
{字段1:值1,字段2:{嵌套字段1:嵌套值1,嵌套字段2:[{},{},{}]}}
和字典很像

在json数据中，字段值必须是以下的数据类型：
字符串
数字
对象（嵌套数据）
数组
布尔值
null #表示空

'''

'''
# json数据在用json()提取之前，类型是字符串
#用json()方法提取以后，会在方法底层经过数据类型的转换，转换称一个对象，为字典或者列表


 
'''

'''
method:请求方法 get post
url：请求网址(数据所在的地址)

headers：请求头字段的关键字参数
cookies：用户身份标识，字典
proxies：ip代理的关键字参数

params:(可选的)请求的查询参数
data:请求的请求参数<post请求>

allow_redirects:是否允许重定向，默认是允许的--就是通过各种方法将各种网络请求重新定个方向转到其它位置
timeout：设置请求响应的时间---时间超过会报错
verify：是否验证证书<ca证书 ssl证书>   如果代码请求的地址没有证书，程序会报错
verify = False 忽视证书验证

json:json提交的请求参数<post请求>
auth：权限认证

'''

# url ='https://image.so.com/i?q=%E9%A3%8E%E6%99%AF&src=srp'


'''
https://image.so.com/i?q=%E9%A3%8E%E6%99%AF&src=srp
查询参数：
    ？前面是请求的地址，后面是请求的一系列查询参数
    &分割每一个查询参数
    
所有的查询参数都是二值型数据
'''
#
'''
import requests
url1 = 'https://image.so.com/i'

params = {
    'q':'%E9%A3%8E%E6%99%AF',
    'src':'srp'
}
#?可加可不加
response = requests.get(url = url1,params=params)
# print(response.request.url)

# url编码问题
#url编码：默认url中不支持中文，默认遇到中文会经过url编码
#url编码构成：百分号+字母+数字

print(requests.utils.quote('风景')) #对中文进行编码
print(requests.utils.unquote('%E9%A3%8E%E6%99%AF'))#对中文进行解码
'''

'''
count = 0
import requests
url = 'http://www.kfc.com.cn/kfccda/ashx/GetStoreList.ashx'

def page_data(page):
    #构建的post请求的请求参数<表单数据>
    data = {
        'cname': '',
        'pid': '',
        'keyword': '武汉',
        'pageIndex': page,
        'pageSize': '10',
    }
    return data
headers ={
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36'

}
#查询参数params-----url的？后面是查询参数
params = {'op': 'keyword'}
#请求数据
#data 关键字传递post请求提交的请求参数
for page in range(1,4):
    data = page_data(page)
    response = requests.post(url = url,headers = headers,data = data,params = params)
    json_data = response.json()

    data1= json_data['Table1']
    # print(data1)

    for data in data1:
        dic = {
            '名字':data['storeName'],
            '所在城市':data['cityName'],
            '地址':data['addressDetail']
        }
        count +=1
        print(count)
        print(dic)
'''

'''
import requests
url = 'http://www.kfc.com.cn/kfccda/ashx/GetStoreList.ashx'

def get_page(name):
    #构建的post请求的请求参数<表单数据>
    data = {
        'cname': '',
        'pid': '',
        'keyword': name,
        'pageIndex': '1',
        'pageSize': '10',
    }
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36'
    }
    params = {'op': 'keyword'}
    response = requests.post(url = url,headers = headers,data = data,params=params)
    json_data = response.json()
    rowcount = json_data['Table'][0]['rowcount']
    # print(rowcount)
    if rowcount%10 >0:
        page = rowcount // 10 + 1
    else:
        page = rowcount // 10

    return page

def send_requests(city_name):
    count = 0
    pages = get_page(city_name)
    for page in range(1,pages+1):
        print(f'正在抓取{city_name} 的第{page}页数据')
        data = {
            'cname': '',
            'pid': '',
            'keyword': city_name,
            'pageIndex': page,
            'pageSize': '10',
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36'
        }
        params = {'op': 'keyword'}
        response = requests.post(url=url, headers=headers, data=data, params=params)
        json_data = response.json()
        data1 = json_data['Table1']
        # print(data1)
        for data in data1:
            dic = {
                '名字': data['storeName'],
                '所在城市': data['cityName'],
                '地址': data['addressDetail']
            }
            count +=1
            print(count)
            print(dic)



city_list = ['北京','武汉','广州']
for city in city_list:
    send_requests(city)

'''
'''
----------------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------------
'''

'''
css选择器：专门用提取HTML中的数据
正则表达式：在小范围的字符串中提取你想要的数据----大范围的字符串需要有良好的正则表达式功底
xpath节点提取：抓门用提取 HTML 中的数据，xml<之前主流的数据交换格式>  json



1、常见的数据类型
结构化数据  ---可以用关系型数据库标识和储存，表现为二维形式的数据
半结构化数据---半关系模型的、有基本固定结构模式的数据，例如日志文件、XML文件、JSON文件
非结构化数据---各种文档、图片、音频/视频，一般储存为二进制的数据格式

2、HTML--超文本标记语言
HTML是一种标记语言，而不是编程语言

html 形式上           head、body、title
'''

'''
import parsel  #第三方模块  数据解析的模块(css选择器、xpath、re)


#1. html 数据<str>转换为一个对象
selector = parsel.Selector()
print(selector)
#Selector 就具有一系列数据解析的方法

#所有通过css选择器解析出来的数据都是对象（Selector）
p = selector.css('p')
print(p)

#get() 从解析的数据中返回第一个数据，直接把字符串数据返回给我们
getall() 从解析的数据中返回所有符合条件的数据

p = selector.css('p').getall()

# 只有标签具有 class 属性才可以用类选择器提取对应的标签
# . 代表提取标签的类型为（class）
# 具有相同类属性的标签都会被提取
# 类选择器就是通过标签的类属性精确定位
result = selector.css('.top').getall()
print(result)

'''

'''
1.类属性中有空格
    空格在css语法中表示取嵌套标签里面的数据
    如果在class类属性中包含了空格，空格需要用.代替
2. 使用类选择器，有可能解析出来的数据量不对：
    可以往上级标签定位，然后向下取
    
'''
'''
id选择器 一般标签的id属性在 html 中是唯一的
# ‘#’ 是用id选择器提取数据
# 只要标签有 id 属性，都会被提取
#contend 代表要提取的标签的 id 属性
result = selector.css('#contend').getall()
print(result)


'''
'''
组合选择器：
各标签、class、id是可以组合使用的

result = selector.css('p#contend.top').getall()

'''

'''
import parsel
selector = parsel.Selector(html)

# :: 表示属性选择器，当我们提取到标签以后，需要获取标签特定的值进行提取(标签包含的文本，标签的属性值)，就可以使用属性选择器
# text 代表提取标签包含的文本（所有符号都是英文形式下的符号）    
result = selector.css('p::text').getall()


::attr(href)   根据标签中包含的属性名字提取属性值
href           a 标签中需要提取属性值的属性名字

selector.css('a::attr(href)').getall()

'''

'''
: 表示伪类选择器--想取满足条件的第几个
#nth-child 满足条件的第几个元素
# :nth-child(2) 表示选择满足标签的第二个元素--类似于索引，但从1开始
#伪类选择器只能选一个，想去一个区间的标签，可以用列表切片

使用方法----------result = selector.css('p:nth-child(2)::text').getall()


'''

'''
css爬虫步骤
1、找数据所对应的地址
2、请求地址数据
3、数据解析
3.1 --转换数据类型
3.2 --解析数据

'''
# import requests
# import parsel
#
# url = 'https://www.guahao.com/expert/61409/%E5%86%85%E7%A7%91'
#
# response = requests.get(url = url)
#
# html_data = response.text
# # print(html_data)
#
# selector = parsel.Selector(html_data)
# selector.css('')
#
# a = 'hello'
# b = 'world'


'''
正则表达式
一个元字符只能匹配一个字符
. 可以匹配除换行符以外的任意字符
re.S 模式匹配，此模式能够让 . 匹配到换行符
\d  匹配数字字符（0-9）
\D  匹配非数字字符
\s  匹配空白字符（\n \t tab ...)
\S  匹配非空白字符
\w  匹配所有的单词字符
\W  匹配所有的非单词字符
+   匹配前一个字符一次或者多次（至少要有一次）
*   匹配前一个字符零次或者多次（至少可以没有）
[]  表示字符集，字符集里面罗列的内容才能匹配，一个字符集只能匹配一个字符串

贪婪模式：正则的默认匹配模式，你的字符串只要满足正则表达式规则，尽可能多的匹配数据

.*  匹配除了换行符以外的任意字符，默认情况下是贪婪模式
.*? 非贪婪模式，匹配一次返回一次
?   匹配1次或者0次

#{start,stop}  表示数量词，限制前一个元字符出现的次数，闭区间
例如：result = re.findall('Super劫Zed: \d{6,11}@qq.com,text)


re模块

re.match(pattern,string,flags)  
从匹配的字符串中，只能从头部开始，根据正则语法匹配

pattern 需要匹配的字符串（正则）
string  要在哪个字符串中匹配
flags   模式匹配

re.compile()  #将正则表达式编译称一个对象

pattern = re.compile('\d+')
result = re.findall(pattern,str1)

\t、\n、\\、...如果出现在需要匹配的字符串中，那么需要转义，把具有特殊含义的字符当作普通的字符串处理
'''

'''
tel = '15171286980'
import re

def func(temp):
    str_ = temp.group()
    return str_[:3] + '****' + str_[-4:]

result = re.sub('\d{11}',func,tel)
print(result)
'''

'''
xpath 语法中：
/ 表示从根节点开始提取（用的很少），还表示取下一级标签----如果打算从根节点开始提取的话，必须从根节点开始提取

// 表示跨结点提取，不用考虑节点的位置
.  表示选取当前节点，使用场景:需要对选取的标签进行二次提取的时候，要用到这个点
.. 表示选取当前节点的父节点--用的很少
@  又两个用途，1、使用@做标签的精确定位  2、可以用@根据属性名字获取属性值
text() 在获取到标签后，可以用此语法获取标签包含的文本

result = selector.xpath('//li/@class|//a/text()').getall()
name = li.xpath('.//*[@class='info']//*[@class='hd']//a//span[1]/text().get()
| 表示多条件查询，所有两边分别是两个条件，满足其中一个条件就会被提取到（逻辑或）
'''

'''
1.文件保存
文件尾缀：是给软件识别 .mp3 
-------------------------------------------
文件写入模式
a 表示 追加--在文件尾部追加数据
w 表示 写入--首先会清空当前文件所有数据，然后写入
r 表示 读取
wb 写入二进制数据--用于图片和视频、音频、字体文件
-------------------------------------------
常见编码方式
GBK windows系统下默认的编码
utf-8 
--------------------------------------------
.read()    读取打开文件的所有数据
readline() 读取文件的一行数据
readlines()读取文件的所有数据，返回一个列表，每一行都是列表里面单独的元素


import os #文件目录操作的模块 内置模块

1.创建文件夹
os.mkdir('python_data') #创建文件夹，括号内部填写文件的路径（绝对路径/相对路径）
                   #当文件夹已存在时，无法创建该文件
2.修改文件夹
os.rename('需要修改名字的文件夹/文件','新名字')
os.rename('python_data','python110')---重命名文件夹/文件

3.删除文件夹
os.rmdir('python110') #删除文件夹，只能删除空文件夹

4.删除文件
os.remove('hello.txt')

5.获取当前目录文件
os.listdir('python_data')---获取当前文件夹（python_data）下面所有的文件，返回一个列表，括号内部指定的是文件的路径

6.判断文件/文件夹是否存在
os.path.exists('python_data') ---查看python文件夹是否存在，如果文件路径存在，返回True，不存在返回False

用法
if not os.path.exists('python_data\\title'):
    os.mkdir('python_data\\title')
    

'''
'''
excel文件保存
import openpyxl

1.创建工作簿对象
work_book = openpyxl.Workbook()

2.创建表<对象>
sheet = work_book.create_sheet('表1')

3.保存此工作簿对象
work_book.save('实例.xlsx')

4.数据写入
sheet['A1'] = 'A1'
sheet['B1'] = 'B1'

cell() 是单元格对象,row表示单元格的行数，column表示单元格的列数
sheet.cell(row=1,column=1).value = '111'
sheet.cell(row=2,column=2).value = '333'

sheet.append()---整行写入数据到表格中，括号内部需要传数据容器(元组，列表)
如果输入对象是元组，则一行一行写入数据
sheet.append((1,2,3,4,5))
sheet.append((6,7,8,9,10))

5.根据excel文加名字读取
work_book = openpyxl.load_workbook('实例.xlsx')

# 获取所有表的名字
print(work_book,sheetnames)

sheet = work_book['表1']

'''
'''
json数据

1.json数据的序列化与反序列化

json数据本身是字符串



import json

data = {
    'name':'ACME',
    'shares':100,
    'price':542023
}

#json 序列化操作：把对象(data为字典)转换成json字符串
json_str = json.dumps(data)

#json 反序列化操作：把json字符串转化成一个对象
json_dict = json.loads(data)

#将对象转化成字符串 就可写入文件
with open('data.json',mode = 'w',encoding = 'utf-8') as f:
    f.write(json.dumps(data))
    
with open('data.json',mode = 'w',encoding = 'utf-8') as f:
    f.read(json.dumps(data))


在默认情况下--json数据的序列化操作如果遇到了中文，会把中文经过unicode编码

data_str = json.dumps(data,ensure_ascii = False) ---不使用unicode编码，然后用with open的时候指定保存为utf-8的编码方式


'''

'''
csv 文件
import csv
---------------------------------------------------------------
ll = [[1,2,3,4],[1,2,3,4],[5,6,7,8],[5,6,7,8]]


#csv写入数据
#针对列表写表头，只能用字符串普通方式写入
with open('data.csv',mode = 'a',encoding = 'utf-8',newline = '') as f:
    f.write('字段1','字段2','字段3','字段4\n')
    csv_write = csv.writer(f)
    for i in ll:
    csv_write.writerow(i)
    
    
#csv读取数据
with open('data.csv',mode = 'a',encoding = 'utf-8',newline = '') as f:
    print(f.read())


newline=''------指定新行，消除csv数据的空行
csv.writer(f)---通过csv模块实例化的数据写入对象，括号内部传递打开的文件对象
writerow(i)----整行写入数据，i 指代的是一个数据容器（元组，列表)

------------------------------------------------------------------
csv写入字典

list_dict = {
            '公司名字': company_name,
            '职位名字': job_name,
            '地区': area,
            '经验': job_exp,
            '学历': job_edu,
            '薪资': providesalary_text,
            '招聘详情页': job_href,
        }
        
with open('dict.csv',mode = 'a',encoding = 'utf-8',newline = '') as f:
    #csv 数据用于字典的特有的一个方法
    #DictWrite
    #    第一个参数就是打开的文件对象
    #    第二个参数就是使用字典写数据的时候，可以根据字典的键作为csv的表头
    fieldnames=['公司名字', '职位名字', '地区', '经验', '学历', '薪资', '招聘详情页']
    csv_write = csv.DictWrite(f,fieldnames = fieldnames)
   
    #写入表头
    csv_write.writeheader()
   
    #写入数据
    for i in list_dict:
        csv_write.writerow(i)
            
       

'''
'''
----------------------------------------------------------------------------
----------------------------------------------------------------------------
'''

'''
requests 高级用法
一.使用python代码维持网页登入状态（下次登入不需要重新输入账号密码）

cookies来源：
    服务器生成：Set—Cookie    服务器后台效验cookies，效验的字段我们目前也不清楚
    浏览器生成：一般服务器不会效验
    js生成cookies：JavaScript  js逆向<解密>

cookies字段
    服务器只会针对特定的字段进行校验，具体是哪一个cookies片段信息需要具体情况具体分析，可以测试
    
    requests模块发送的请求是单次请求，从第二次开始，cookies信息就会被清空
    状态保持：请求需要又用户状态<注册>

1.cookies保证同一用户
session----------------模拟用户登入，模拟用户验证、注册等需要用到session，保证验证状态

session = requests.Session()----创建一个会话维持对象，可以直接用此对象发送多次请求，多次的请求具有状态的维持

第一次使用session
response =session.get(url = url,headers = headers,cookies = cookies)

第二次使用session
register_response = session.post(url = register_url,json = json_data)----因为使用session，url被更新，headers不用再加，cookies也是

使用session会话维持，如果你要打印cookies字段信息，只能在第一次打印


二.resquests 异常处理

1.异常处理----------try expect
-----------------------------------------------------------------------
2.encoding 表示设置对象的编码 apparent——encoding 自动识别响应体的编码

1.  response.encoding = response.apparent_encoding-----自动识别
2.  response.encoding = 'utf-8'-----精确指定

utf-8   保存数据，如果仍然有乱码，可以考虑使用utf-8-sig（针对常见的表情符号和表情文字）   
gbk     保存数据，如果仍然有乱码，可以考虑使用gb2312

'''

'''
一、selenium快速上手
----------------------------------------------
#导入浏览器模块
from selenium import webdriver

#1.创建一个浏览器对象
driver = webdriver.Chrome(executable_path='chromedriver.exe')

#2.执行浏览器自动化操作
driver.get('https://www.baidu.com/')

#3.阻止浏览器关闭 输入函数
driver.close() #退出当前页面
driver.quit()  #退出整个浏览器
----------------------------------------------
二、driver对象的常用方法

1.driver.get()---------通过浏览器对象指定请求网址
2.driver.save_screenshot()----------截取浏览器页面的图片，括号内部指定的是图片的名字
3.driver.page_source  (属性)-------查看页面渲染以后的数据(html)
****用selenium和用浏览器在elements中看到的数据有可能有出入****-----做数据解析一定要以获取到的代码为准
4.driver.get_cookies()--------------查看请求以后的cookies------用此方法可以得到加密后的cookies
5.driver.current_url--------------查看当前请求的url地址
6.driver.driver.maximize_window()-----最大化浏览器
7.driver.minimize_window()-----最小化浏览器
---------------------------------------------------------------------------
三、元素提取---解析数据

find_element_by_css_selector()
find_element_by_xpath_selector()

#在selenium中，css、xpath语法不支持属性提取、文本提取，只能定位标签对象
-------------------------------------------------------------
text
    方法，可以提取标签包含的文本内容，支持链式调用
    因为css/xpath语法不支持属性提取，只能定位标签
-------------------------------------------------------------
    find_element_by_css_selector().text
    find_element_by_xpath_selector().text
-------------------------------------------------------------
get_attribute()
    方法，可以根据属性名获取标签的属性值
    find_element_by_css_selector().get_attribute('href')-----------获取html中的href属性值
---------------------------------------------------------------------------------------
四、标签对象的动作(输入/点击)
1.找到你要操作的标签对象
input_label = find_element_by_css_selector('.inp input')
input_label.send_keys('长津湖')----------向输入框中输入数据，支持链式调用
2.点击查询标记
click_label = driver.find_element_by_css_selector('.bn input')---click()  执行标签的点击操作，需要标签具有点击事件
click_label.click()
---------------------------------------------------------------------------------------
五、等待页面渲染
1.强制等待
time.sleep(5)
#动态页面<弹出标签元素，弹出验证码>  
2.隐式等待
driver.implicitly_wait(5)
#智能化等待
#打开很多页面，只需要设置一次隐式等待，后续的页面都会遵循隐式等待的规则
#如果超过了隐式等待的事件，仍然会报错
---------------------------------------------------------------------------------------
#driver.window_handles  获取当前浏览器所有出窗口，返回一个窗口对象的列表
#可以根据窗口对象的列表索引切换窗口
driver.switch_to.window(driver.window_handles[1])
----------------------------------------------------------------------------------------------------------

'''

'''
selenium 进阶

一、嵌套网页
1.嵌套网页的标签一般是<iframe>----------使用selenium采集数据时，一般ctrl + F4 查询是否有iframe标签
获取网页html---首先进入嵌套网页----通过代码进入

第一种进入嵌套网页的形式：
    driver.switch_to.frame(0)
    print(driver.page_source)
1.根据索引进入到嵌套网页，索引值时从0考试的 2.索引值不能超出当前已有的嵌套数量
---------------------------------------------------------------------
第二种进入嵌套网页的形式：(推荐)
    iframe =driver.find_element_by_css_selector('#g_iframe')
    driver.switch_to.frame(iframe)----进入到嵌套网页
    driver.switch_to.parent_frame()----从当前嵌套网页，切换回父网页
#根据嵌套网页的<iframe>标签进入到嵌套网页

--------------------------------------------------------------------- 

二、执行javaScript代码
#js滚动页面的代码
    document.documentElement.scrollTop  指定页面滚动条的位置
    document.documentElement.scrollHeight  获取当前页面最大高度的代码
    js = 'document.documentElement.scrollTop = 1000'-----页面滚动下拉1000像素
    js = 'document.documentElement.scrollTop = document.documentElement.scrollHeight'---直接下拉到页面底部

driver.execute_script(js)------------执行js代码的方法，括号内部传递js代码

-----------------------------------------------------------------------------

三、鼠标动作链

from selenium import webdriver
from selenium.webdriver import ActionChains  #导入鼠标动作链的功能

driver = webdriver.Chrome(executable_path='chromedriver.exe')
driver.get('https://www.runoob.com/try/try.php?filename=jqueryui-api=droppable')

#进入嵌套网页
driver.switch_to_frame(0)
#找到可以拖动标签
drag = driver.find_element_by_css_selector('#draggable')
#找到放置的位置
drop = driver.find_element_by_css_selector('#droppable')

#鼠标动作链执行---------支持链式调用
#1.创建一个动作链对象
action = ActionChains(driver)

#2.定义一个动作,此动作到目前还没有被执行
action.drag_and_drop(drag,drop)

#3.执行动作链 perfrom()
action.perform()--释放
action.click()--点击
-------------------------------------------
#导入浏览器模块
from selenium import webdriver
from selenium.webdriver import ActionChains
driver = webdriver.Chrome(executable_path='chromedriver.exe')
driver.get('https://gitee.com/')

driver.find_element_by_css_selector('.item.git-nav-user__login-item').click()
time.sleep(2)
driver.find_element_by_css_selector('.login-password__account-input').send_keys('15171286980')
time.sleep(2)
driver.find_element_by_css_selector('#user_password').send_keys('hzw1234567')
time.sleep(2)
driver.find_element_by_css_selector('.ui.fluid.orange.submit.button.large:nth-child(1)').click()

time.sleep(2)
#登入后解决消息提示<弹出（耗时）>
ActionChains(driver).move_by_offset(1073,661).click().perform()
--------------------------------------------------------------------------------------------------
三、selenium设置无头模式

from selenium.webdriver.chrome.options import Options #浏览器的可选项
from selenium import webdriver

chrome_options = Options()  #实例化对象
chrome_options.add_argument('--headless') #添加一个可选项 #--headless 指代无头模式
#创建浏览器对象的时候添加配置
driver = webdriver.Chrome(executable_path='chromedriver.exe',chrome_options = chrome_options)
driver.get('https://www.baidu.com')
print(driver.page_source)

driver.quit()

1.无头模式一般在项目完成后添加
2.无头模式对于一些浏览器的动作操作，不支持
-------------------------------------------------------------
无头模式，方法2：

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
driver = webdriver.Chrome(executable_path='chromedriver.exe', options=chrome_options)
driver.get('https://music.163.com/#/discover/toplist?id=3778678')
-----------------------------------------------------
四、selenium应对检测

from selenium import webdriver
#创建浏览器对象的时候添加配置
driver = webdriver.Chrome(executable_path='chromedriver.exe')
------------------------------------------------------
#配置浏览器的某个属性
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",{
    "source":"""
    Object.defineProperty(navigator,'webdriver',{
    get: () => undefined
    })
"""   
})
------------------------------------------------------
driver.get('')
print(driver.page_source)
driver.quit()


'''

# import requests
#
# requests = requests.get('https://www.baidu.com/')
# print(requests)
#
# print(requests.text)

'''
爬虫项目的实现步骤

1.找数据所在的链接地址  <分析网页性质：静态网页/动态网页>
2.代码模拟请求地址数据
3.数据提取，提取需要的数据，剔除不需要的数据
4.数据保存(本地、数据库）

'''
# import requests
#
# url = 'http://www.ibiqu.org/138_138753/174248505.html'
# response = requests.get(url)
# print(response)



