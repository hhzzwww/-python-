"""
1. 程序启动，显示信息管理系统欢迎界面，并显示功能菜单
2. 用户用数字选择不同的功能
3. 根据功能选择，执行不同的功能
4. 需要记录学生的 **姓名**、**语文成绩**、**数学成绩**、**英语成绩** 、**总分**
5. 如果查询到指定的学生信息，用户可以选择 **修改** 或者 **删除** 信息
6. 进入或退出时加载或保存数据
"""
from openpyxl import load_workbook, Workbook

message = """**************************************************
欢迎使用【学生信息管理系统】V1.0
请选择你想要进行的操作
1. 新建学生信息
2. 显示全部信息
3. 查询学生信息
4. 修改学生信息
5. 删除学生信息

0. 退出系统
**************************************************"""


# 程序启动的时候从本地加载数据
def load_data_form_excel():
    students = []
    wb = load_workbook(filename='sample.xlsx')
    sheet = wb.active

    for row in sheet.rows:
        data = [col.value for col in row]
        students.append({'name': data[0], 'math': data[1], 'chinese': data[2], 'english': data[3]})
    return students


# 程序结束的时候将诗句保存到本地

def save_data_to_excel(students):
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    for student in students:
        data = list(student.values())
        ws.append(data)

    # Save the file
    wb.save("sample.xlsx")


students = load_data_form_excel()


def create_student():
    """
    新建一个学生信息对象
    :return: student
    """
    name = input('请输入学生的姓名：')
    chinese = input('请输入学生的语文成绩：')
    math = input('请输入学生的数学成绩：')
    english = input('请输入学生的英语成绩：')
    student = {'name': name, 'math': int(math), 'chinese': int(chinese), 'english': int(english)}
    return student


def show_students(students):
    print('姓名	    语文	    数学	    英语	    总分')
    for student in students:
        print(student['name'], student['chinese'], student['math'], student['english'],
              student['chinese'] + student['math'] + student['english'], sep='\t\t')


def search_student(name):
    """
    根据名字查询学生信息
    :return: student or False
    """
    for student in students:
        if name == student['name']:
            return student
    return False


while True:
    print(message)
    action = input('请选择你想要的进行的操作：')
    if action == '1':
        student = create_student()
        students.append(student)

    elif action == '2':
        show_students(students)
    elif action == '3':
        name = input('请输入需要查询的学生名字：')
        student = search_student(name)
        if student is False:
            print(name, ' 学员不存在')
        else:
            show_students([student])
    elif action == '4':
        name = input('请输入需要修改的学生姓名：')

        student = search_student(name)
        if not (student is False):
            new_name = input('请重新输入学生的名字：')
            math = input('请重新输入数学成绩：')
            chinese = input('请重新输入语文成绩：')
            english = input('请重新输入英语成绩：')
            if math != '':
                student['math'] = int(math)
            if chinese != '':
                student['chinese'] = int(chinese)
            if english != '':
                student['english'] = int(english)
            if new_name != '':
                student['name'] = new_name
        else:
            print(name, ' 此学员不存在 无法修改')
    elif action == '5':
        name = input('请输入需要删除的学生姓名：')

        student = search_student(name)
        if student:
            students.remove(student)
        else:
            print(name, ' 此学员不存在 无法修改')
    elif action == '0':
        save_data_to_excel(students)
        print('退出系统成功，欢迎下次光临')
        break
    else:
        print('请选择正确的操作')

# 语义化
